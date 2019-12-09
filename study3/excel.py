from openpyxl import load_workbook

work_book = load_workbook('prac01.xlsx')
work_sheet = work_book['prac']

# 데이터를 읽어봅니다.
#temp_text = work_sheet.cell(row = 1, column = 1).value

for i in range(1,10):
    temp_text = work_sheet.cell(row=i, column=1).value
    print(temp_text)
