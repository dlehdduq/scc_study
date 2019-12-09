import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

work_book = load_workbook('prac03.xlsx')
work_sheet = work_book['prac']

# 타겟 URL을 읽어서 HTML를 받아오고,
# 헤더로 가상의 유저를 만들어서 이 유저가 해당 URL 접근하는 것 처럼 보이게 한다!
headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)

# HTML을 BeautifulSoup이라는 라이브러리를 활용해 검색하기 용이한 상태로 만듦
# soup이라는 변수에 "파싱 용이해진 html"이 담긴 상태가 됨
# 이제 코딩을 통해 필요한 부분을 추출하면 된다.
soup = BeautifulSoup(data.text, 'html.parser')

#############################
# (입맛에 맞게 코딩)

# select를 이용해서, tr들을 불러오기
movies = soup.select('#old_content > table > tbody > tr')
row = 1;
# movies (tr들) 의 반복문을 돌리기

work_sheet.delete_rows(1)
work_sheet.delete_rows(2)

#work_sheet.append(["랭킹","제목","평점"])
work_sheet.cell(row=row, column=1, value="랭킹")
work_sheet.cell(row=row, column=2, value="제목")
work_sheet.cell(row=row, column=3, value="평점")
#work_sheet.cell(row=row, column=4, value="")

for movie in movies:
    # movie 안에 a 가 있으면,
    rank = movie.select_one('td.ac ')
    title = movie.select_one('td.title > div > a')
    point = movie.select_one('td.point')
    if title is not None:
        # a의 text를 찍어본다.
        #print (rank('img')[0]['alt'],title.text,point.text)
        work_sheet.cell(row=row, column=1, value=rank('img')[0]['alt'])
        work_sheet.cell(row=row, column=2, value=title.text)
        work_sheet.cell(row=row, column=3, value=point.text)
    row += 1;
work_book.save('prac03.xlsx')